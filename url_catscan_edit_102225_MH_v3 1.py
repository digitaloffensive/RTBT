import requests
import time
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter    

import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Zscaler's block page often contains a unique HTML element or a specific text string.
# You must update this variable with a unique string found on your client's Zscaler
# block page to ensure the script can accurately identify blocked sites.
# For example, it might be something like "Access Denied by Zscaler" or "Your organization has blocked this site."
ZSCALER_BLOCK_SIGNATURE = "you don't have permission to visit this site"
ZSCALER_BLOCK_SIGNATURE = "request forbidden by administrative rules"

# A dictionary of URL categories to test, with one or more sample URLs for each.
URL_CATEGORIES_TO_TEST = {
    "ENTERTAINMENT": ["www.hollywood.com", "www.sydneyoperahouse.com"],
    "MUSIC": ["www.itunes.com", "www.spotify.com"],
    "OTHER_ENTERTAINMENT_AND_RECREATION": ["rockcreekrvpark.com", "destinywateradventures.com"],
    "RADIO_STATIONS": ["radioducinema.com", "statsradio.com"],
    "TELEVISION_AND_MOVIES": ["imdb.com", "movies.com"],
    "STREAMING_MEDIA": ["netflix.com", "hotstar.com"],
    "NEWS_AND_MEDIA": ["www.cnn.com", "www.asiaone.com"],
    "CLASSIFIEDS": ["www.st701.com", "www.gumtree.com"],
    "CORPORATE_MARKETING": ["www.billabong.com", "www.crocs.com"],
    "FINANCE": ["www.hsbc.com", "www.citibank.com"],
    "INSURANCE": ["www.tataaig.com", "www.unitedhealthgroup.com"],
    "TRADING_BROKARAGE_INSURANCE": ["ameritrade.com", "etrade.com"],
    "OTHER_BUSINESS_AND_ECONOMY": ["etrind.com.br", "dfsafrica.org"],
    "PROFESSIONAL_SERVICES": ["www.micematters.com", "www.issworld.com"],
    "CONTINUING_EDUCATION_COLLEGES": ["www.duke.edu", "www.ox.ac.uk"],
    "HISTORY": ["www.hyperhistory.com", "www.besthistorysites.net"],
    "K_12": ["www.k12.com", "www.ilacademy.net"],
    "OTHER_EDUCATION": ["www.writework.com"],
    "REFERENCE_SITES": ["en.wikipedia.org", "www.dictionary.com"],
    "SCIENCE_AND_TECHNOLOGY": ["www.scitechdaily.com", "www.livescience.com"],
    "WEB_BANNERS": ["www.buysellads.com", "www.fusionads.net"],
    "AI_ML_APP": ["www.openai.com", "bard.google.com"],
    "CDN": ["cdn.espn.com", "bighost.be"],
    "DEVELOPER_TOOLS": ["rubyonrails.org", "github.com"],
    "DNS_OVER_HTTPS": ["dns.google/dns-query", "cloudflare-dns.com/dns-query"],
    "FILE_CONVERTORS": ["www.freepdfconvert.com", "www.freeconvert.com"],
    "FILE_HOST": ["www.yousendit.com", "www.rapidshare.com"],
    "GENERAL_AI_ML": ["aimagazine.com", "www.aitrends.com"],
    "IMAGE_HOST": ["www.flickr.com", "www.kodakgallery.com"],
    "OSS_UPDATES": ["swcdn.apple.com", "windowsupdate.microsoft.com"],
    "OTHER_INFORMATION_TECHNOLOGY": ["kwonnam.pe.kr", "bellsouthemailsettings.com"],
    "PORTALS": ["qq.com", "naver.com"],
    "SAFE_SEARCH_ENGINE": ["www.safesearchkids.com", "fragfinn.de"],
    "SHAREWARE_DOWNLOAD": ["www.download.cnet.com", "www.filehippo.com"],
    "TRANSLATORS": ["translate.google.com", "www.freetranslation.com"],
    "WEB_HOST": ["www.siteground.com", "www.myown.eu"],
    "WEB_SEARCH": ["www.google.com", "www.yahoo.com"],
    "BLOG": ["www.blogger.com", "onlinejournalismblog.com"],
    "DISCUSSION_FORUMS": ["www.cellar.org", "www.wsc-forum.de"],
    "INTERNET_SERVICES": ["www.singnet.com.sg", "www.imagin.com"],
    "ONLINE_CHAT": ["finnchat.com", "msngr.com"],
    "OTHER_INTERNET_COMMUNICATION": ["softyupdates.com", "ipixo.com"],
    "P2P_COMMUNICATION": ["www.bittorrent.com", "www.limewire.com"],
    "REMOTE_ACCESS": ["www.teamviewer.com", "www.logmein.com"],
    "WEB_CONFERENCING": ["www.webex.com", "www.zoom.us"],
    "EMAIL_HOST": ["mail.google.com", "www.hotmail.com"],
    "JOB_SEARCH": ["www.monster.com", "www.kellyservices.com"],
    "GLOBAL_INT_OFC365_ALLOW": ["smtp.office365.com"],
    "GLOBAL_INT_OFC365_DEFAULT": ["ssw.live.com", "storage.live.com"],
    "GLOBAL_INT_OFC365_OPTIMIZE": ["outlook.office.com", "outlook.office365.com"],
    "GOVERNMENT": ["www.usa.gov", "www.gov.sg"],
    "MILITARY": ["www.af.mil", "www.army.mil"],
    "OTHER_GOVERNMENT_AND_POLITICS": ["wec24.org", "covid-sb.org"],
    "POLITICS": ["www.nrsc.org", "www.pap.org.sg"],
    "TRAVEL": ["www.contiki.com", "www.singaporeair.com"],
    "VEHICLES": ["www.toyota.com", "www.volkswagen.com"],
    "ADULT_SEX_EDUCATION": ["www.sexedlibrary.org", "www.itsyoursexlife.com"],
    "ADULT_THEMES": ["www.truveo.com", "www.singlebrides.com"],
    "SEXUALITY": ["kicktattoo.com", "tattoogrid.net"],
    "K_12_SEX_EDUCATION": ["www.teensource.org", "www.sexetc.org"],
    "LINGERIE_BIKINI": ["www.wacoal.com", "www.victoriassecret.com"],
    "NUDITY": ["nakedsipandpaint.com", "prostitutkimoskvyxxx.com"],
    "OTHER_ADULT_MATERIAL": ["teamcrazy.za.net", "smilemakerscollection.com"],
    "PORNOGRAPHY": ["www.playboy.com", "www.sex.com"],
    "SOCIAL_ADULT": ["www.okcupid.com", "www.ashleymadison.com"],
    "MARIJUANA": ["www.leafly.com", "thcbiomed.com"],
    "DRUGS": ["livwell.com", "buyecstasy.com"],
    "GAMBLING": ["www.casino.com", "www.singaporepools.com.sg"],
    "ANONYMIZER": ["www.proxyway.com", "www.your-freedom.net"],
    "COMPUTER_HACKING": ["www.astalavista.net", "www.cracks.am"],
    "COPYRIGHT_INFRINGEMENT": ["sci-hub.tw", "bingemachine.com"],
    "MATURE_HUMOR": ["thejokeyard.com", "badmovies.org"],
    "OTHER_ILLEGAL_OR_QUESTIONABLE": ["salesreceiptstore.com", "katcr.co"],
    "PROFANITY": ["www.babyfight.com", "www.blowwboston.com"],
    "QUESTIONABLE": ["www.churchofsatan.org", "www.satanicchurch.com"],
    "MILITANCY_HATE_AND_EXTREMISM": ["www.newnation.org", "www.klanparenthood.com"],
    "TASTELESS": ["www.morticom.com", "www.gore2gasm.com"],
    "VIOLENCE": ["stranakrovi.com", "serienkillers.de"],
    "WEAPONS_AND_BOMBS": ["dsparmory.co", "nragungiveaway.org"],
    "OTHER_GAMES": ["www.worldofwarcraft.com", "agame.com"],
    "SOCIAL_NETWORKING_GAMES": ["www.facebook.com/FarmVille", "www.facebook.com/WSOP"],
    "HEALTH": ["multiformelegym.com", "ochsstaywell.com"],
    "ALT_NEW_AGE": ["thetarotguide.com", "mirsularii.com"],
    "CULT": ["theflatearthsociety.org", "medaglia-miracolosa.it"],
    "OTHER_RELIGION": ["stmarkstn.org", "qcforjesus.com"],
    "TRADITIONAL_RELIGION": ["www.chc.org.sg", "www.buddhanet.net"],
    "ONLINE_AUCTIONS": ["www.ebay.com", "www.onlineauction.com"],
    "SPECIALIZED_SHOPPING": ["www.amazon.com"],
    "OTHER_SHOPPING_AND_AUCTIONS": ["thecholmeleyarms.co.uk", "shoppiego.com"],
    "REAL_ESTATE": ["www.redas.com", "www.propnex.com"],
    "FAMILY_ISSUES": ["maritallaws.com", "resetting-the-family.com"],
    "OTHER_SOCIAL_AND_FAMILY_ISSUES": ["hopecle.org", "familiesforfamilies.net"],
    "SOCIAL_ISSUES": ["kandoo.me", "robindiangelo.com"],
    "ALCOHOL_TOBACCO": ["www.martell.com", "melbournehookah.com.au"],
    "ART_CULTURE": ["www.artandculture.com", "www.metmuseum.org"],
    "DINING_AND_RESTAURANT": ["www.mcdonalds.com", "www.hungrygowhere.com"],
    "HOBBIES_AND_LEISURE": ["jenniemasterson.com", "takemefishing.org"],
    "LIFESTYLE": ["www.lambda.org", "clubmask.com"],
    "OTHER_SOCIETY_AND_LIFESTYLE": ["carolineandmichael2020.com", "ashlandandblake.com"],
    "SOCIAL_NETWORKING": ["www.facebook.com", "www.friendster.com"],
    "SPECIAL_INTERESTS": ["www.greenpeace.org", "www.audi-denkwerkstatt.de"],
    "SPORTS": ["espn.go.com", "www.nba.com"],
    "ENCR_WEB_CONTENT": ["mask.icloud.com", "app.stupendo.co"],
    "DYNAMIC_DNS": ["dyndns.com", "no-ip.com"],
    "OTHER_SECURITY": ["trustwave.ctscloud.com", "microsoftinternetsafety.net"],
    "ADWARE_OR_SPYWARE": ["www.spywareremove.com", "www.virusspy.com"],
    "ENTERTAINMENT_OR_RECREATION": ["www.hollywood.com", "sydneyoperahouse.com"]
}

def check_category_access(category_name, urls):
    """
    #Tests access to a given URL category by attempting to fetch URLs.
    
    Args:
        category_name (str): The name of the URL category.
        urls (list): A list of URLs to test for the category.
    
    Returns:
        str: "Allowed" if at least one URL is reachable, "Blocked" otherwise.
    """

    print(f"--- Testing {category_name} ---")
    allowed_count = 0
    blocked_count = 0
    timeout_count = 0
    url_details = []

    # Use a session to handle cookies and persistent headers
    session = requests.Session()
    # Common browser headers
    browser_headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
        'Accept-Language': 'en-US,en;q=0.9',
        'Accept-Encoding': 'gzip, deflate, br',
        'Connection': 'keep-alive',
        'Upgrade-Insecure-Requests': '1',
        'Cache-Control': 'max-age=0',
        # 'Referer': 'https://www.google.com/' # Optional, can add if needed
    }
    session.headers.update(browser_headers)

    for url in urls:
        if not url.startswith(('http://', 'https://')):
            url = f"http://{url}"

        status = ""
        status_code = None
        notes = ""
        try:
            response = session.get(url, timeout=10, verify=False)
            status_code = response.status_code
            if ZSCALER_BLOCK_SIGNATURE.lower() in response.text.lower():
                status = "Blocked"
                notes = "Zscaler block signature detected"
                blocked_count += 1
            elif response.status_code == 200:
                status = "Allowed"
                notes = "Status code: 200 OK"
                allowed_count += 1
            else:
                if 300 <= response.status_code < 400:
                    meaning = "Redirect (the resource has moved or requires further action)"
                elif 400 <= response.status_code < 500:
                    meaning = "Client Error (the request was invalid or unauthorized)"
                elif 500 <= response.status_code < 600:
                    meaning = "Server Error (the server failed to fulfill a valid request)"
                else:
                    meaning = "Other/Unknown response code"
                status = "Possibly Blocked"
                notes = f"Status code: {response.status_code} - {meaning}"
                blocked_count += 1
        except (requests.exceptions.Timeout, requests.exceptions.ConnectionError) as e:
            status = "Timeout / Connection Error"
            notes = f"Network issue or timeout: {e}"
            timeout_count += 1
        except requests.exceptions.RequestException as e:
            status = "Blocked"
            notes = f"Connection error or timeout: {e}"
            blocked_count += 1

        url_details.append({
            "category": category_name,
            "url": url,
            "status": status,
            "status_code": status_code,
            "notes": notes
        })
        time.sleep(1)

    return {
        "Allowed": allowed_count,
        "Blocked": blocked_count,
        "Timeout / Connection Error": timeout_count,
        "Total": len(urls),
        "details": url_details
    }

def main():
    """
    Main function to run the Zscaler URL category test program.
    """
    print("Starting Zscaler URL Category Test...")
    print("\n")
    
    results = {}
    all_details = []
    for category, urls in URL_CATEGORIES_TO_TEST.items():
        result = check_category_access(category, urls)
        results[category] = result
        all_details.extend(result["details"])
        print("\n")

    # Print a summary of the results
    print("Test Complete. Summary of Results:")
    print("\n")
    for category, summary in results.items():
        print(f"{category:<35}: Allowed: {summary['Allowed']}, Blocked: {summary['Blocked']}, Timed Out: {summary['Timeout / Connection Error']}, Total: {summary['Total']}")

    # Export to Excel
    export_results_to_excel(all_details, filename="zscaler_url_results.xlsx")

def export_results_to_excel(details, filename="zscaler_url_results.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "URL Test Results"

    # Define headers
    headers = ["Index", "Category", "URL", "Status", "Status Code", "Notes"]
    ws.append(headers)

    # Bold header row
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Add borders to all cells
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_idx = 2
    for idx, entry in enumerate(details, 1):
        ws.append([
            idx,
            entry["category"],
            entry["url"],
            entry["status"],
            entry["status_code"],
            entry.get("notes", "")
        ])
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).border = border
        row_idx += 1

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(filename)
    print(f"Results exported to {filename}")

if __name__ == "__main__":
    main()
